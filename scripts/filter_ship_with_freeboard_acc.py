"""
批量干舷法计算干舷精度验证
"""
import open3d as o3d
import numpy as np
import os
import json
import open3d.visualization.gui as gui
import ast
from scipy.spatial import cKDTree
from scipy.optimize import minimize
import glob
import pandas as pd
from scipy.stats import entropy
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from tqdm import tqdm
from sklearn.decomposition import PCA


def voxel_filter(points, voxel_size=0.5):
    """
    点云体素滤波。

    :param points: 点云数据，形状为 (n, 3) 的 NumPy 数组。
    :type points: np.ndarray
    :param voxel_size: 体素大小，默认为 0.5。
    :type voxel_size: float
    :return: 体素滤波后的点云。
    :rtype: np.ndarray
    """
    pcd = o3d.geometry.PointCloud()
    pcd.points = o3d.utility.Vector3dVector(points)
    down_pcd = pcd.voxel_down_sample(voxel_size=voxel_size)
    return np.asarray(down_pcd.points)

def select_center(points, return_index=False):
    """
    取船重心部分
    :param points: 点云
    :return: 筛选后的点云
    """

    # 将点云投射到xoy平面上
    xoy_points = points.copy()
    xoy_points[:, 2] = 0
    min_bound = xoy_points.min(axis=0)
    max_bound = xoy_points.max(axis=0)
    range = abs(max_bound[0] - min_bound[0]) / 6
    point_center  = (min_bound + max_bound) / 2
    # 筛选 X 轴在均值 ±20 范围内的点
    x_filter_indices = np.where((points[:, 0] >= (point_center[0] - range)) & (points[:, 0] <= (point_center[0] + range)))[0]
    process_points = points[x_filter_indices]
    if return_index:
        return process_points, x_filter_indices
    else:
        return process_points

def statistical_outlier(points, nb_neighbors=30, std_ratio=1.0, is_inliers=False):
    """统计滤波去除离群点"""
    # 构建 KDTree
    tree = cKDTree(points)

    # 计算每个点的平均距离
    distances, _ = tree.query(points, k=nb_neighbors)
    mean_distances = np.mean(distances, axis=1)

    # 计算标准差
    mean_distance = np.mean(mean_distances)
    stddev = np.std(mean_distances)

    # 找到离群点
    inliers = np.abs(mean_distances - mean_distance) < std_ratio * stddev
    if is_inliers:
        return inliers
    else:
        return points[inliers]

def compute_principal_direction(points):
    mean = np.mean(points, axis=0)
    centered = points - mean
    cov = np.cov(centered, rowvar=False)
    vals, vecs = np.linalg.eigh(cov)
    principal = vecs[:, np.argmax(vals)]
    if np.dot(principal, [1, 0, 0]) < 0:
        principal = -principal
    return principal

def build_rotation_matrix(principal_direction):
    """
    构建将主方向对齐到 x 轴的旋转矩阵，仅绕 y 轴和 z 轴旋转
    :param principal_direction: 主方向向量
    :return: 旋转矩阵
    """
    # 归一化主方向向量
    principal_direction = principal_direction / np.linalg.norm(principal_direction)
    # 计算绕 y 轴的旋转角度
    angle_y = np.arctan2(principal_direction[2], principal_direction[0])
    # 构建绕 y 轴的旋转矩阵
    rotation_matrix_y = np.array([
        [np.cos(angle_y), 0, np.sin(angle_y)],
        [0, 1, 0],
        [-np.sin(angle_y), 0, np.cos(angle_y)]
    ])
    # 计算绕 z 轴的旋转角度
    # 先将主方向向量绕 y 轴旋转到 xz 平面上
    rotated_principal_direction = np.dot(rotation_matrix_y, principal_direction)
    angle_z = np.arctan2(rotated_principal_direction[1], rotated_principal_direction[0])
    # 构建绕 z 轴的旋转矩阵
    rotation_matrix_z = np.array([
        [np.cos(angle_z), np.sin(angle_z), 0],
        [-np.sin(angle_z), np.cos(angle_z), 0],
        [0, 0, 1]
    ])
    # 组合旋转矩阵
    rotation_matrix = np.dot(rotation_matrix_z, rotation_matrix_y)

    return rotation_matrix


def compute_average_density(points, k_neighbors=8):
    if points.size == 0:
        return 0.0
    kdtree = cKDTree(points)
    distances, indices = kdtree.query(points, k=k_neighbors)
    avg_density = k_neighbors / np.mean(distances[:, -1])
    return avg_density

def compute_roughness_entropy(points, k_neighbors=20, radius=1.0):
    # ---------- taipu_main_side. 预处理 ----------
    points = np.asarray(points, dtype=np.float64)
    mask = np.isfinite(points).all(axis=1)
    points = points[mask]
    points = np.unique(points, axis=0)
    if points.shape[0] == 0:
        raise ValueError("输入点云为空或全为 NaN/Inf！")
    # ---------- 2. cKD-Tree ----------
    kdtree = cKDTree(points)
    # ---------- 3. 粗糙度（新：点到局部平面距离的标准差） ----------
    n_points = points.shape[0]
    roughness = np.zeros(n_points)
    # 查询邻域点
    distances, indices = kdtree.query(points, k=k_neighbors)
    # 计算每个点的局部粗糙度
    for i in range(n_points):
        neigh = points[indices[i]]  # K×3
        centroid = neigh.mean(axis=0)  # 3
        centered_neigh = neigh - centroid
        U, S, Vt = np.linalg.svd(centered_neigh)
        normal = Vt[-1]  # 最小特征值对应的主成分

        # 点到平面距离 = |(p - p0)·n|
        distances = np.abs(np.dot(centered_neigh, normal))
        roughness[i] = np.std(distances)
    # ---------- 4. 熵 ----------
    total = roughness.sum()
    if total == 0:
        return 0.0
    probs = roughness / total
    probs = probs[probs > 0]
    roughness_entropy = entropy(probs, base=2)

    return roughness_entropy


def apply_rotation(points, center_points, density, points_thresh, density_thresh, mode=None):
    """
    应用旋转矩阵到点云（优化版 - 保持原方法不变）
    :param points: 点云
    :return:
    """
    if mode == "bridge":
        if len(points) < points_thresh:
            xoy_points = center_points.copy()
        else:
            xoy_points = points.copy()
    else:
        xoy_points = points.copy()
    # 将点云投射到xoy平面上
    xoy_points[:, 2] = 0
    unique_xoy_points = np.unique(xoy_points, axis=0)

    if len(points) < points_thresh and density < density_thresh:  # 15000有可能也是一半船
        # 将船按几何中心分成两部分
        min_bound = unique_xoy_points.min(axis=0)
        max_bound = unique_xoy_points.max(axis=0)
        xoy_center = (min_bound + max_bound) / 2
        xoy_lower = unique_xoy_points[unique_xoy_points[:, 0] < xoy_center[0]]
        xoy_higher = unique_xoy_points[unique_xoy_points[:, 0] >= xoy_center[0]]
        width_lower = xoy_lower[:, 1].max() - xoy_lower[:, 1].min()
        width_higher = xoy_higher[:, 1].max() - xoy_higher[:, 1].min()
        if width_lower < width_higher:
            selected_part = xoy_lower
        else:
            selected_part = xoy_higher

    else:
        selected_part = xoy_points

    principal_direction = compute_principal_direction(selected_part)

    # 3. 构建旋转矩阵并应用
    rotation_matrix = build_rotation_matrix(principal_direction)
    center = np.mean(center_points, axis=0)
    rotated_points = np.dot(center_points - center, rotation_matrix.T) + center
    rotation_total_points = np.dot(points - center, rotation_matrix.T) + center

    return rotation_total_points, rotated_points, rotation_matrix, center

def reverse_rotation(rotated_points, rotation_matrix, center, offset=0.0):
    """
    将旋转后的点反变换回原始坐标系
    :param rotated_points: 旋转后的点云
    :param rotation_matrix: 旋转矩阵
    :param center: 旋转中心（保持不变）
    :param offset: 事先人为加上的 y 方向平移量
    :return: 反变换后的点云
    """
    # taipu_main_side. 先减掉“抬高后”的旋转中心
    shifted_points = rotated_points - (center + [0, offset, 0])  # 用临时抬高后的中心
    # 2. 逆旋转
    reversed_points = np.dot(shifted_points, rotation_matrix)
    # 3. 加回“原始”旋转中心（不动）
    original_points = reversed_points + (center + [0, offset, 0])

    return original_points

def gaussian(x, A, u, sigma):
    return A * np.exp(-0.5 * ((x - u) / sigma) ** 2)

def f(x, A, u_list, sigma):
    return np.sum([gaussian(x, A, u, sigma) for u in u_list])

def gaussian_approximation(u_list, sigma=0.25, distinguishing_coefficient=10, A=1):
    if len(u_list) == 0:
        return None
    select_value = list(range(int((min(u_list) - 1)*distinguishing_coefficient), int((max(u_list) + 1)*distinguishing_coefficient)))
    assert len(select_value) > 0
    max_value_list = []
    for i in select_value:
        result = minimize(lambda x: -f(x, A, u_list, sigma), x0=i / distinguishing_coefficient,method='Powell')
        max_value = -result.fun
        max_value_list.append(max_value)
    init_value =select_value[max_value_list.index(max(max_value_list))] / distinguishing_coefficient
    result = minimize(lambda x: -f(x, A, u_list, sigma), x0=init_value,method='Powell')

    # 输出结果
    max_x = result.x[0]
    return max_x

def divide_points(points, axis=0, n=3):
    # 根据指定轴坐标对点云进行排序，并获取排序后的索引
    sorted_indices = np.argsort(points[:, axis])
    sorted_points = points[sorted_indices]

    # 计算分位点的索引
    total_points = len(points)
    if total_points < n:
        raise ValueError(f"点云中点的数量不足以分成 {n} 等分")

    # 计算每个分组的大小
    group_size = total_points // n
    remainder = total_points % n

    # 分组
    groups = []
    start_index = 0
    for i in range(n):
        end_index = start_index + group_size
        group = sorted_points[start_index:end_index]
        groups.append(group)
        start_index = end_index

    # 如果有余数，将余数单独分成一组
    if remainder > 0:
        group = sorted_points[start_index:]
        groups.append(group)

    return groups

def find_most_frequent_value(arr, precision=1, n=2, threshold=0.1):
    # 将数组中的每个元素保留到指定的小数位数
    rounded_arr = np.round(arr, precision)

    # 对四舍五入后的值进行排序
    sorted_values = np.sort(rounded_arr)

    # 合并相近的值（前后相差不到 0.taipu_main_side）
    groups = []
    current_group = []

    for value in sorted_values:
        if not current_group:
            current_group.append(value)
        else:
            # 检查当前值与组内所有值的误差是否都在 0.taipu_main_side 以内
            if all(abs(value - v) <= threshold for v in current_group):
                current_group.append(value)
            else:
                groups.append(current_group)
                current_group = [value]

    if current_group:
        groups.append(current_group)

    # 找到最大的组
    largest_group = max(groups, key=len)

    # 最大组占比不到 20 % 或绝对个数不到 3 个，都认为“不集中”
    if len(largest_group) < max(3, n // 5):
        return None

    # 找到所有与众数相对应的原始值
    mask = np.isin(np.round(arr, precision), largest_group)
    corresponding_values = arr[mask]

    # 计算这些值的平均数
    mean_value = np.mean(corresponding_values) if len(corresponding_values) > 0 else None

    return mean_value

def filter_by_y(points, total_points, mode=None, save_ratio=3.0):
    """
    按 y 值滤波：
    taipu_main_side. 计算 y_min, y_max，取中点作为几何中心 y_center
    2. 计算 y 方向宽度 width = y_max - y_min
    3. 保留 |y - y_center| > width / 3 的点
    返回：过滤后的点云 (M, 3)
    """

    if points.ndim != 2 or points.shape[1] != 3:
        raise ValueError("points must be a (N, 3) array")

    y = points[:, 1]
    y_total = total_points[:, 1]
    y_min, y_max = y_total.min(), y_total.max()
    y_center = (y_min + y_max) / 2.0
    width = (y_max - y_min) / 2

    # 计算 y - y_center
    y_diff = y - y_center
    mask = np.abs(y_diff) > (width / save_ratio)

    if mode == 'bank':  # 岸基模式
        mask_1 = mask & (y_diff < 0)
        return points[mask_1], 1
    elif mode == 'bridge':  # 桥基模式
        pos_cnt = np.sum(y > 0)
        if pos_cnt > 0.75 * len(y):  # 大部分点在 y 轴正方向
            mask_1 = mask & (y_diff < 0)  # 靠近坐标轴的一侧
            mask_2 = mask & (y_diff > 0)  # 远离坐标轴的一侧
            if abs(np.sum(mask_1) - np.sum(mask_2)) > 200:
                process_points, direction = (points[mask_2], -1) if np.sum(mask_1) < np.sum(mask_2) else (
                    points[mask_1], 1)
            else:
                process_points, direction = points[mask_1], 1
        else:  # 大部分点在 y 轴负方向
            mask_1 = mask & (y_diff > 0)  # 靠近坐标轴的一侧
            # mask_2 = mask & (y_diff < 0)  # 远离坐标轴的一侧
            # process_points, direction = (points[mask_2], 1) if np.sum(mask_1) < np.sum(mask_2) else (points[mask_1], -1)
            process_points, direction = (points[mask_1], -1)
        return process_points, direction
    else:
        return points[mask], 1

def estimate_normals(points, radius, max_nn):
    """
    自定义法向量估计函数，使用 Open3D 的 KDTreeSearchParamHybrid。
    :param points: 点云数据，形状为 (N, 3) 的 numpy 数组。
    :param radius: 搜索半径。
    :param max_nn: 最大邻近点数。
    :return: 法向量，形状为 (N, 3) 的 numpy 数组。
    """
    # 将点云数据转换为 Open3D 的 PointCloud 对象
    pcd = o3d.geometry.PointCloud()
    pcd.points = o3d.utility.Vector3dVector(points)
    # 构建 KDTree
    pcd_tree = o3d.geometry.KDTreeFlann(pcd)
    # 初始化法向量数组
    normals = np.zeros_like(points)
    # 遍历每个点，计算其法向量
    for i in range(points.shape[0]):
        # 查询第 i 个点的邻近点
        [_, idx, _] = pcd_tree.search_hybrid_vector_3d(points[i], radius, max_nn)
        # 获取邻域点
        neighbors = np.asarray(pcd.points)[idx]
        # 计算邻域点的均值
        mean = np.mean(neighbors, axis=0)
        # 计算协方差矩阵
        centered_points = neighbors - mean
        covariance = np.dot(centered_points.T, centered_points) / len(idx)
        # 计算协方差矩阵的特征值和特征向量
        eigenvalues, eigenvectors = np.linalg.eigh(covariance)
        # 选择最小特征值对应的特征向量作为法向量
        normals[i] = eigenvectors[:, np.argmin(eigenvalues)]

    return normals

def detect_boundaries(points, radius=0.1, k=16, angle_threshold=np.pi * 0.7):
    """
    检测边界点
    :param points: 点云数据，形状为 (N, 3)
    :param k: 每个点的最近邻点数量
    :param angle_threshold: 角度阈值
    :return: 边界点
    """
    # 初始化边界点标记数组
    pcd = o3d.geometry.PointCloud()
    pcd.points = o3d.utility.Vector3dVector(points)
    boundary_points = np.zeros(len(points), dtype=bool)
    # 计算法向量
    normals = estimate_normals(points, radius=radius, max_nn=k)
    pcd.normals = o3d.utility.Vector3dVector(normals)
    pcd.orient_normals_consistent_tangent_plane(k=k)
    normals = np.asarray(pcd.normals)
    # 遍历每个点，检查其是否为边界点
    kdtree = o3d.geometry.KDTreeFlann(pcd)
    for i in range(len(points)):
        [_, idx, _] = kdtree.search_knn_vector_3d(points[i], k)  # 查找16个最近邻点
        neighbor_normals = normals[idx]

        # 计算当前点与邻居点的法向量夹角
        angles = np.arccos(np.clip(np.dot(normals[i], neighbor_normals.T), -1.0, 1.0))
        if np.any(angles > angle_threshold):  # 角度大于0.6π的点被认为是边界点
            boundary_points[i] = True

    return points[~boundary_points]

def mahalanobis_filter_vec(points, nb_neighbors=16, std_ratio=2.0):
    """
    针对甲板平面优化的马氏距离滤波（向量化实现）。

    Parameters
    ----------
    points : np.ndarray
        输入点云，形状 (N, 3)。
    nb_neighbors : int
        每个点的邻居数量（不含自身）。
    std_ratio : float
        阈值倍数，inlier < mean + std_ratio * std。

    Returns
    -------
    inlier_pts : np.ndarray
        滤波后的点云，形状 (K, 3)。
    mask : np.ndarray
        保留掩码，形状 (N,)，bool。
    """
    tree = cKDTree(points)
    n = points.shape[0]

    # 一次性找邻居 (N, nb_neighbors)
    dists, idx = tree.query(points, k=nb_neighbors + 1)
    idx = idx[:, 1:]  # 去掉自身
    neigh = points[idx]  # (N, nb_neighbors, 3)

    # 邻居均值 (N, 3)
    μ = neigh.mean(axis=1)  # (N, 3)

    # 邻居中心化 (N, nb_neighbors, 3)
    diff = neigh - μ[:, None, :]  # (N, nb_neighbors, 3)

    # 批量协方差 (N, 3, 3)
    cov = np.einsum('nki,nkj->nij', diff, diff) / nb_neighbors
    cov += 1e-6 * np.eye(3)  # 正则化

    # 批量逆矩阵 (N, 3, 3)
    inv_cov = np.linalg.inv(cov)

    # 自身到邻居均值的残差 (N, 3)
    self_diff = points - μ  # (N, 3)

    # 马氏距离平方 (N,)
    md2 = np.einsum('ni,nij,nj->n', self_diff, inv_cov, self_diff)
    md = np.sqrt(md2)

    # 阈值判定
    mean_md = md.mean()
    std_md = md.std()
    mask = md < (mean_md + std_ratio * std_md)

    return points[mask], mask

def apply_rotation_matrix_to_point_cloud(points, R):
    """
    对点云中的每个点应用旋转矩阵
    """
    rotated_points = (R @ points.T).T  # 对每个点应用旋转矩阵
    return rotated_points

class LidarFunc(object):
    def __init__(self, parser_dict, **kwargs):
        self.parser_dict = parser_dict
        self.extra_params = kwargs

    def freeboard_calculation(self, points, water_level=None):
        # points = apply_rotation_matrix_to_point_cloud(points, np.asarray(
        #     [[0.9999936784236516, 4.441341249212944e-07, 0.00355571547472536],
        #      [4.441341249212944e-07, 0.9999999687965295, -0.0002498134158006251],
        #      [-0.00355571547472536, 0.0002498134158006251, 0.9999936472201811]]))
        # points = points[points[:, 2] > -11.38]
        o3d.utility.random.seed((42))
        density_all = compute_average_density(points, k_neighbors=16)
        points = voxel_filter(points, voxel_size=self.parser_dict["voxel_size"] * density_all)
        x_min = points[:, 0].min()
        x_max = points[:, 0].max()
        x_center = (x_max + x_min) / 2
        if x_center < self.parser_dict['x_min'] or x_center > self.parser_dict['x_max']:
            return x_center, None, None, None, None, None
        center_points = select_center(points)
        density = compute_average_density(center_points, k_neighbors=16)
        (
            rotation_total_points,
            rotation_center_points,
            rotation_matrix,
            center,  # 旋转中心
        ) = apply_rotation(points, center_points, density,
                           points_thresh=self.parser_dict['points_thresh'],
                           density_thresh=self.parser_dict['density_thresh'],
                           mode=self.parser_dict['func_mode'])
        if self.parser_dict['func_mode'] == "bridge":
            boundary_points = detect_boundaries(rotation_center_points, radius=0.5 * density, k=16,
                                                angle_threshold=np.pi * 0.7)
            process_points, direction = filter_by_y(boundary_points, rotation_total_points,
                                                    mode=self.parser_dict['func_mode'])
        elif self.parser_dict['func_mode'] == "bank":
            process_points, direction = filter_by_y(rotation_center_points, rotation_total_points,
                                                    mode=self.parser_dict['func_mode'],
                                                    save_ratio=self.parser_dict['save_ratio'])
        if len(process_points) > self.parser_dict['filter_thresh']:
            nb_neighbors = self.parser_dict['nb_neighbors'] * density
            std_ratio = self.parser_dict['std_ratio'] * density
            process_points, _ = mahalanobis_filter_vec(process_points, nb_neighbors=nb_neighbors, std_ratio=std_ratio)
        n = max(1, int(len(process_points) / self.parser_dict['divide_thresh']))
        divide_rotation_center_list = divide_points(process_points, axis=0, n=n)
        freeboard_upper_points = []
        for divide_rotation_center in divide_rotation_center_list:
            # 对每个部分提取y坐标
            y_coords = divide_rotation_center[:, 1] + self.parser_dict['axis_offset']
            score = direction * np.abs(y_coords)
            order = np.argsort(score)
            # 2. 逐步扩大窗口，直到方差超标
            for k in range(1, len(order) + 1):
                idx_win = order[:k]
                var = np.var(y_coords[idx_win])
                if var >= self.parser_dict['var_thresh']:
                    # 超标了，取上一次的结果
                    selected = order[:k - 1]
                    break
            else:
                # 整个数组方差都不超标，全取
                selected = order
            y_top_index = selected

            # 对其z轴大小排序
            max_z_idx_in_top = np.argmax(divide_rotation_center[y_top_index][:, 2])
            global_max_z_idx = y_top_index[max_z_idx_in_top]
            freeboard_upper_point = divide_rotation_center[global_max_z_idx]  # 获取该点的 z 坐标作为干舷的上界
            freeboard_upper_points.append(freeboard_upper_point)
         # 拿出freeboard_upper_points保留两位小数后频数最多的点作为freeboard_upper_point
        freeboard_upper_points = np.asarray(freeboard_upper_points)
        freeboard_upper_points = reverse_rotation(freeboard_upper_points, rotation_matrix, center)
        freeboard_upper_bound = find_most_frequent_value(freeboard_upper_points[:, 2], n=n)
        if freeboard_upper_bound is None:
            freeboard_upper_bound = np.median(freeboard_upper_points[:, 2])
        if water_level is not None:
            freeboard_lower_bound = water_level
            deck = abs(freeboard_upper_bound - freeboard_lower_bound)
        else:
            freeboard_lower_bound = center_points[:, 2].min()
            deck = freeboard_upper_bound - freeboard_lower_bound
        return (
                x_center,
                self.parser_dict["voxel_size"] * density_all,
                density,
                freeboard_upper_bound,
                freeboard_lower_bound,
                deck,
        )


def remove_outliers_iqr(measurements, factor=1.5, variance_threshold=1e-3):
    """
    基于四分位数（IQR）方法去除测量值中的异常值。

    参数:
        measurements (numpy.ndarray): 测量值数组。
        factor (float): IQR的倍数，默认为1.5。

    返回:
        tuple:
            - filtered_measurements (numpy.ndarray): 去除异常值后的测量值数组。
            - outliers (numpy.ndarray): 被识别为异常值的数组。
    """
    # 确保 measurements 是 NumPy 数组
    measurements = np.array(measurements)
    # 将 None 值替换为 np.nan
    measurements = np.where(measurements == None, np.nan, measurements)
    if np.var(measurements) < variance_threshold:
        # 过滤掉 NaN 值
        measurements = measurements[~np.isnan(measurements)]
        return measurements  # 方差太小，跳过 IQR 去噪

    # 将数组转换为数值类型
    try:
        measurements = pd.to_numeric(measurements, errors='coerce')
    except ValueError as e:
        raise ValueError("数据中包含无法转换为数值的值，请检查输入数据。") from e

    # 过滤掉 NaN 值
    measurements = measurements[~np.isnan(measurements)]

    return measurements

def split_filename_from_path(file_path):
    """
    从完整路径中提取文件名（不包括扩展名），并按 '_' 分隔为多个部分。
    参数: file_path (str): 完整的文件路径。
    返回: list: 文件名的各个部分。
    """
    # 提取文件名（包括扩展名）
    file_name_with_extension = os.path.basename(file_path)
    # 去掉扩展名
    file_name, _ = os.path.splitext(file_name_with_extension)
    # 使用 split() 方法按 '_' 分隔文件名
    parts = file_name.split('_')
    # 去除空字符串部分（如果存在）
    parts = [part for part in parts if part]

    return parts

def read_txt_dict(txt_data_path):
    # 获取所有txt文件
    file_names = glob.glob(f'{txt_data_path}/*.txt', recursive=True)
    for i, file_name in enumerate(file_names):
        print(f"[{i}]-{file_name}")
    file_idx = int(input("选择筛选的txt文件>>"))

    try:
        # 打开文件并逐行读取
        with open(file_names[file_idx], 'r', encoding='utf-8') as file:
            lines = file.readlines()  # 读取所有行

        # 用于存储解析后的字典
        txt_dicts = []

        # 遍历每一行，尝试解析为字典
        for line in lines:
            line = line.strip()  # 去掉首尾空格和换行符
            if not line:
                continue  # 跳过空行
            try:
                # 替换单引号为双引号（如果需要）
                line = line.replace("'", '"')
                # 使用 ast.literal_eval 安全地解析为字典
                txt_dict = ast.literal_eval(line)
                txt_dicts.append(txt_dict)  # 将解析后的字典添加到列表中
            except SyntaxError as e:
                print(f"解析错误：{e}，内容为：{line}")
            except Exception as e:
                print(f"其他错误：{e}，内容为：{line}")

    except FileNotFoundError:
        print(f"文件未找到，请检查路径：{file_names[file_idx]}")
    except Exception as e:
        print(f"发生错误：{e}")
    return txt_dicts

def find_lidar_interval_by_input(input_value, txt_dicts):
    # 遍历txt_dicts中的每一项
    for txt_item in txt_dicts:
        # 检查local_path是否包含输入值
        if input_value in txt_item.get("local_path", ""):
            # 返回lidar_interval对应的值
            return txt_item.get("lidar_interval")
    # 如果没有匹配到，返回None
    return None

def draw_steady_png(data, steady_value, channel_name, save_path):
    y = np.asarray(data, dtype=float)

    # 空数组 & 等值保护
    if y.size == 0:
        y = np.array([0.0])
    if np.all(y == y[0]):
        y = y + 1e-6 * np.random.randn(len(y))

    idx   = np.arange(len(y))
    total = len(y)

    # 稳态判定（最长区间）
    win   = max(5, min(30, len(y) // 4))
    roll  = pd.Series(y).rolling(win, min_periods=1).std()
    thresh = roll.quantile(0.25) if len(y) > win else np.inf
    steady_mask = roll < thresh

    # 提取所有区间，只保留最长一段
    starts, ends = [], []
    in_steady = False
    for i, val in enumerate(steady_mask):
        if val and not in_steady:
            starts.append(i)
            in_steady = True
        elif not val and in_steady:
            ends.append(i)
            in_steady = False
    if in_steady:
        ends.append(len(y))

    # 最长区间
    if starts:
        lengths = np.array(ends) - np.array(starts)
        max_idx = np.argmax(lengths)
        start, end = starts[max_idx], ends[max_idx]
        n_steady = end - start
    else:
        start, end, n_steady = None, None, 0

    # ±0.taipu_main_side 区间
    if steady_value is not None:
        mask_02 = (y >= steady_value - 0.1) & (y <= steady_value + 0.1)
        n_02    = mask_02.sum()
        closest_idx = np.argmin(np.abs(y - steady_value))
        closest_val = y[closest_idx]
    else:
        n_02 = 0
        closest_idx = None
        closest_val = None

    # 绘图
    fig, ax = plt.subplots(figsize=(10, 5))
    fig.subplots_adjust(left=0.06, right=0.75, top=0.9, bottom=0.15)

    ax.set_title(f'Extracting Steady-State Value from {channel_name}', fontsize=11, pad=8)
    ax.plot(idx, y, color='#0072bd', lw=1.2)
    ax.scatter(idx, y, color='#0072bd', s=12, zorder=3)

    # 仅画最长稳态条带
    if start is not None:
        ax.axvspan(start, end, color='#ff7f0e', alpha=0.25)

    if steady_value is not None:
        ax.axhspan(steady_value - 0.1, steady_value + 0.1,
                   color='yellow', alpha=0.25)
        ax.axhline(steady_value, color='red', ls='--', lw=1.5)
        ax.scatter(closest_idx, closest_val, color='darkgreen', s=50, zorder=5)

    # 竖排图例
    legend_texts = [
        'Measured data',
        f'Total Points : {total}',
        f'Steady Region : {n_steady}',
        f'Steady ±0.2 : {n_02}',
        f'Steady Value : {steady_value:.2f}' if steady_value is not None else 'Steady Value : N/A',
        f'Closest Point : {closest_val:.2f}' if closest_val is not None else 'Closest Point : N/A'
    ]
    ax.legend(legend_texts, loc='best', fontsize=8, frameon=False)

    y_min, y_max = y.min(), y.max()
    if y_min == y_max:
        y_min, y_max = y_min - 1, y_max + 1
    ax.set_xlim(0, max(1, len(y)))
    ax.set_ylim(y_min - 0.05 * (y_max - y_min), y_max + 0.05 * (y_max - y_min))

    ax.set_xlabel('Index', fontsize=9)
    ax.set_ylabel('Value', fontsize=9)
    ax.grid(True, which='both', ls='--', lw=0.5, alpha=0.3)
    if os.path.exists(save_path):
        try:
            os.remove(save_path)
        except PermissionError:
            print(f"文件被占用：{save_path}")

    plt.savefig(save_path, dpi=200, bbox_inches='tight')
    plt.close(fig)

if __name__ == '__main__':
    o3d.utility.random.seed(42)
    directory_path    = r"../data/WuHanQingShanDaQiao/data_0716_0718"
    output_excel_path = './table/qingshan.xlsx'
    png_dir_path = './table'
    folder_name = ".qingshan"

    with open('json_files/parser_dict_bridge.json', 'r', encoding='utf-8') as file:
        data = json.load(file)
    parser_dict = data.get('parser_dict', {})
    # 初始化雷达方法
    lidar_func = LidarFunc(parser_dict)

    # 断点续跑（仅按 Excel 最后一行）
    if os.path.exists(output_excel_path):
        last = pd.read_excel(output_excel_path).iloc[-1]
        last_dir = last['Directory']
    else:
        last_dir = last_subdir = None

    dirs = [d for d in os.listdir(directory_path)
            if os.path.isdir(os.path.join(directory_path, d))]
    start_index = dirs.index(last_dir) + 1 if last_dir else 0

    for dir in dirs[start_index:]:
        detail_dir = os.path.join(directory_path, dir, "detail")
        current_path = os.path.join(detail_dir, "deck_pcd")
        pcd_files = glob.glob(os.path.join(current_path, '*.pcd'))

        frame_freeboard_upper_bound = []
        loop = tqdm(pcd_files, desc=f"{dir}/deck_pcd")
        for pcd_file in loop:
            l = split_filename_from_path(pcd_file)
            # pcd = o3d.io.read_point_cloud(pcd_file, format="pcd")
            # # 其他过滤
            # pcd.points = o3d.utility.Vector3dVector(np.asarray(pcd.points)[: int(l[-1])])
            # (x_center, voxel_size, density,
            #  freeboard_upper_bound, _, _) = lidar_func.freeboard_calculation(np.asarray(pcd.points))
            try:
                pcd = o3d.io.read_point_cloud(pcd_file, format="pcd")
                # 其他过滤
                pcd.points = o3d.utility.Vector3dVector(np.asarray(pcd.points)[: int(l[-1])])
                (x_center, voxel_size, density,
                 freeboard_upper_bound, _, _) = lidar_func.freeboard_calculation(np.asarray(pcd.points))
                loop.set_description(f'{dir}')
                loop.set_postfix(x_center=x_center, voxel_size=voxel_size, density=density)
            except Exception:
                freeboard_upper_bound = None
            if freeboard_upper_bound is not None:
                frame_freeboard_upper_bound.append(freeboard_upper_bound)

            # ---------- 计算高斯峰值 ----------
        if not frame_freeboard_upper_bound:
            steady_state_freeboard_upper_bound = None
        else:
            steady_state_freeboard_upper_bound = gaussian_approximation(
                remove_outliers_iqr(frame_freeboard_upper_bound))

            # ---------- 画稳态图 ----------
        png_name = f"steady.png"
        png_path = os.path.join(png_dir_path, folder_name, dir)
        os.makedirs(png_path, exist_ok=True)
        png_path = os.path.join(png_path, png_name)  # 放在 dir 目录下
        os.makedirs(os.path.dirname(png_path), exist_ok=True)
        draw_steady_png(remove_outliers_iqr(frame_freeboard_upper_bound),
                        steady_state_freeboard_upper_bound,
                        f'{dir}',
                        png_path)

        # 追加写入 Excel（无空行）
        data = [{
            'Directory': dir,
            'Steady State Freeboard Upper Bound': steady_state_freeboard_upper_bound,
            'Steady State Plot': f'=HYPERLINK("{os.path.abspath(png_path)}", "Plot")'
        }]
        df = pd.DataFrame(data)

        if os.path.exists(output_excel_path):
            wb = load_workbook(output_excel_path)
            ws = wb.active
            next_row = ws.max_row + 1
            for col, val in enumerate(df.iloc[0], 1):
                ws.cell(row=next_row, column=col, value=val)
            wb.save(output_excel_path)
        else:
            df.to_excel(output_excel_path, index=False)

        print(f"峰值 {steady_state_freeboard_upper_bound}")

    print(f"全部完成，结果见 {output_excel_path}")