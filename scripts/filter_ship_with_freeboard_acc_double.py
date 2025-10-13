"""
批量双向干舷法干舷精度验证
"""
import open3d as o3d
import numpy as np
import os
import open3d.visualization.gui as gui
import ast
from scipy.spatial import KDTree
from scipy.optimize import minimize
import glob
import pandas as pd
from scipy.stats import entropy
from collections import Counter
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from tqdm import tqdm

def select_center(points, return_index=False):
    """
    取船重心部分
    :param points: 点云
    :return: 筛选后的点云
    """

    # 将点云投射到xoy平面上
    xoy_points = points.copy()
    xoy_points[:, 2] = 0
    min_bound = xoy_points.min(axis=0)
    max_bound = xoy_points.max(axis=0)
    range = abs(max_bound[0] - min_bound[0]) / 6
    point_center  = (min_bound + max_bound) / 2
    # 筛选 X 轴在均值 ±20 范围内的点
    x_filter_indices = np.where((points[:, 0] >= (point_center[0] - range)) & (points[:, 0] <= (point_center[0] + range)))[0]
    process_points = points[x_filter_indices]
    if return_index:
        return process_points, x_filter_indices
    else:
        return process_points

def statistical_outlier(points, nb_neighbors=16, std_ratio=1.0, is_inliers=False):
    """统计滤波去除离群点"""
    # 构建 KDTree
    # 去掉含 NaN/inf 的点
    mask = np.isfinite(points).all(axis=1)
    points = points[mask]
    tree = KDTree(points)

    # 计算每个点的平均距离
    distances, _ = tree.query(points, k=nb_neighbors)
    mean_distances = np.mean(distances, axis=1)

    # 计算标准差
    mean_distance = np.mean(mean_distances)
    stddev = np.std(mean_distances)

    # 找到离群点
    inliers = np.abs(mean_distances - mean_distance) < std_ratio * stddev
    if is_inliers:
        return inliers
    else:
        return points[inliers]


def compute_principal_direction(points):
    mean = np.mean(points, axis=0)
    centered = points - mean
    cov = np.cov(centered, rowvar=False)
    vals, vecs = np.linalg.eigh(cov)
    principal = vecs[:, np.argmax(vals)]
    if np.dot(principal, [1, 0, 0]) < 0:
        principal = -principal
    return principal

def build_rotation_matrix(principal_direction):
    """
    构建将主方向对齐到 x 轴的旋转矩阵，仅绕 y 轴和 z 轴旋转
    :param principal_direction: 主方向向量
    :return: 旋转矩阵
    """
    # 归一化主方向向量
    principal_direction = principal_direction / np.linalg.norm(principal_direction)
    # 计算绕 y 轴的旋转角度
    angle_y = np.arctan2(principal_direction[2], principal_direction[0])
    # 构建绕 y 轴的旋转矩阵
    rotation_matrix_y = np.array([
        [np.cos(angle_y), 0, np.sin(angle_y)],
        [0, 1, 0],
        [-np.sin(angle_y), 0, np.cos(angle_y)]
    ])
    # 计算绕 z 轴的旋转角度
    # 先将主方向向量绕 y 轴旋转到 xz 平面上
    rotated_principal_direction = np.dot(rotation_matrix_y, principal_direction)
    angle_z = np.arctan2(rotated_principal_direction[1], rotated_principal_direction[0])
    # 构建绕 z 轴的旋转矩阵
    rotation_matrix_z = np.array([
        [np.cos(angle_z), np.sin(angle_z), 0],
        [-np.sin(angle_z), np.cos(angle_z), 0],
        [0, 0, 1]
    ])
    # 组合旋转矩阵
    rotation_matrix = np.dot(rotation_matrix_z, rotation_matrix_y)

    return rotation_matrix


def compute_average_density(points):
    if points.size == 0:
        return 0.0
    min_b, max_b = points.min(axis=0), points.max(axis=0)
    vol = np.prod(max_b - min_b)
    return len(points) / vol if vol > 1e-12 else float('inf')

def estimate_normals(points, radius, max_nn):
    """
    自定义法向量估计函数，使用 Open3D 的 KDTreeSearchParamHybrid。
    :param points: 点云数据，形状为 (N, 3) 的 numpy 数组。
    :param radius: 搜索半径。
    :param max_nn: 最大邻近点数。
    :return: 法向量，形状为 (N, 3) 的 numpy 数组。
    """
    # 将点云数据转换为 Open3D 的 PointCloud 对象
    pcd = o3d.geometry.PointCloud()
    pcd.points = o3d.utility.Vector3dVector(points)
    # 构建 KDTree
    pcd_tree = o3d.geometry.KDTreeFlann(pcd)
    # 初始化法向量数组
    normals = np.zeros_like(points)
    # 遍历每个点，计算其法向量
    for i in range(points.shape[0]):
        # 查询第 i 个点的邻近点
        [_, idx, _] = pcd_tree.search_hybrid_vector_3d(points[i], radius, max_nn)
        # 获取邻域点
        neighbors = np.asarray(pcd.points)[idx]
        # 计算邻域点的均值
        mean = np.mean(neighbors, axis=0)
        # 计算协方差矩阵
        centered_points = neighbors - mean
        covariance = np.dot(centered_points.T, centered_points) / len(idx)
        # 计算协方差矩阵的特征值和特征向量
        eigenvalues, eigenvectors = np.linalg.eigh(covariance)
        # 选择最小特征值对应的特征向量作为法向量
        normals[i] = eigenvectors[:, np.argmin(eigenvalues)]

    return normals

def compute_roughness_entropy(points,
                              k_neighbors=8,
                              radius=1.0):
    # ---------- taipu_main_side. 预处理 ----------
    points = np.asarray(points, dtype=np.float64)
    # 去 NaN/Inf
    mask = np.isfinite(points).all(axis=1)
    points = points[mask]
    # 去重
    points = np.unique(points, axis=0)
    if points.shape[0] == 0:
        raise ValueError("输入点云为空或全为 NaN/Inf！")
    pcd = o3d.geometry.PointCloud()
    pcd.points = o3d.utility.Vector3dVector(points)
    # ---------- 3. 法向量估计 ----------
    normals = estimate_normals(points, radius=radius, max_nn=k_neighbors)
    normals = np.asarray(normals)

    # ---------- 4. 计算粗糙度 ----------
    roughness = np.zeros(len(pcd.points))
    kdtree = o3d.geometry.KDTreeFlann(pcd)

    for i in range(len(pcd.points)):
        [_, idx, _] = kdtree.search_knn_vector_3d(pcd.points[i], k_neighbors)
        if len(idx) < 3:          # 邻居太少，跳过
            continue
        # 用所有法向量方向的总体标准差作为粗糙度
        roughness[i] = np.linalg.norm(np.std(normals[idx, :], axis=0))

    # ---------- 5. 计算熵 ----------
    total = roughness.sum()
    if total == 0:
        return 0.0
    probs = roughness / total
    probs = probs[probs > 0]        # 避免 log(0)
    roughness_entropy = entropy(probs, base=2)
    return roughness_entropy

def apply_rotation(points, center_points, density):
    """
    应用旋转矩阵到点云（优化版 - 保持原方法不变）
    :param points: 点云
    :return:
    """
    xoy_points = points.copy()
    # 将点云投射到xoy平面上
    xoy_points[:, 2] = 0
    unique_xoy_points = np.unique(xoy_points, axis=0)

    if len(points) < 15000 and density < 0.8:  # 15000有可能也是一半船
        # 将船按几何中心分成两部分
        min_bound = unique_xoy_points.min(axis=0)
        max_bound = unique_xoy_points.max(axis=0)
        xoy_center = (min_bound + max_bound) / 2
        xoy_lower = unique_xoy_points[unique_xoy_points[:, 0] < xoy_center[0]]
        xoy_higher = unique_xoy_points[unique_xoy_points[:, 0] >= xoy_center[0]]
        density_lower = compute_average_density(xoy_lower)
        density_higher = compute_average_density(xoy_higher)
        # 根据密度大小选择部分
        if density_lower < density_higher:
            selected_part = xoy_lower
        else:
            selected_part = xoy_higher

    else:
        selected_part = xoy_points

    principal_direction = compute_principal_direction(selected_part)

    # 3. 构建旋转矩阵并应用
    rotation_matrix = build_rotation_matrix(principal_direction)
    center = np.mean(center_points, axis=0)
    rotated_points = np.dot(center_points - center, rotation_matrix.T) + center

    return rotated_points, rotation_matrix, center

def reverse_rotation(rotated_points, rotation_matrix, center, offset=0.0):
    """
    将旋转后的点反变换回原始坐标系
    :param rotated_points: 旋转后的点云
    :param rotation_matrix: 旋转矩阵
    :param center: 旋转中心（保持不变）
    :param offset: 事先人为加上的 y 方向平移量
    :return: 反变换后的点云
    """
    # taipu_main_side. 先减掉“抬高后”的旋转中心
    shifted_points = rotated_points - (center + [0, offset, 0])  # 用临时抬高后的中心
    # 2. 逆旋转
    reversed_points = np.dot(shifted_points, rotation_matrix)
    # 3. 加回“原始”旋转中心（不动）
    original_points = reversed_points + (center + [0, offset, 0])

    return original_points

def gaussian(x, A, u, sigma):
    return A * np.exp(-0.5 * ((x - u) / sigma) ** 2)

def f(x, A, u_list, sigma):
    return np.sum([gaussian(x, A, u, sigma) for u in u_list])

def gaussian_approximation(u_list, sigma=0.25, distinguishing_coefficient=10, A=1):
    if len(u_list) == 0:
        return None
    select_value = list(range(int((min(u_list) - 1)*distinguishing_coefficient), int((max(u_list) + 1)*distinguishing_coefficient)))
    assert len(select_value) > 0
    max_value_list = []
    for i in select_value:
        result = minimize(lambda x: -f(x, A, u_list, sigma), x0=i / distinguishing_coefficient,method='Powell')
        max_value = -result.fun
        max_value_list.append(max_value)
    init_value =select_value[max_value_list.index(max(max_value_list))] / distinguishing_coefficient
    result = minimize(lambda x: -f(x, A, u_list, sigma), x0=init_value,method='Powell')

    # 输出结果
    max_x = result.x[0]
    return max_x

def divide_points(points, axis=0, n=3):
    # 根据指定轴坐标对点云进行排序，并获取排序后的索引
    sorted_indices = np.argsort(points[:, axis])
    sorted_points = points[sorted_indices]

    # 计算分位点的索引
    total_points = len(points)
    if total_points < n:
        raise ValueError(f"点云中点的数量不足以分成 {n} 等分")

    # 计算每个分组的大小
    group_size = total_points // n
    remainder = total_points % n

    # 分组
    groups = []
    start_index = 0
    for i in range(n):
        # 如果有余数，将余数均匀分配到前面的组中
        end_index = start_index + group_size + (1 if i < remainder else 0)
        group = sorted_points[start_index:end_index]
        groups.append(group)
        start_index = end_index

    return groups


def find_most_frequent_value(arr, precision=1, n=1):
    # 将数组中的每个元素保留到指定的小数位数
    rounded_arr = np.round(arr, precision)

    # 使用 Counter 统计每个元素的出现次数
    value_counts = Counter(rounded_arr)

    # 找到出现次数最多的元素
    most_common_value, _ = value_counts.most_common(1)[0]

    # 找到所有与众数相对应的原始值
    mask = np.round(arr, precision) == most_common_value
    corresponding_values = arr[mask]
    if len(corresponding_values) < n / 5:
        return None

    # 计算这些值的平均值
    mean_value = np.mean(corresponding_values) if len(corresponding_values) > 0 else None

    return mean_value

def filter_by_y(points):
    """
    按 y 值滤波：
    taipu_main_side. 计算 y_min, y_max，取中点作为几何中心 y_center
    2. 计算 y 方向宽度 width = y_max - y_min
    3. 保留 |y - y_center| > width / 3 的点
    返回：过滤后的点云 (M, 3)
    """
    if points.ndim != 2 or points.shape[1] != 3:
        raise ValueError("points must be a (N, 3) array")

    y = points[:, 1]
    y_min, y_max = y.min(), y.max()
    y_center = (y_min + y_max) / 2.0
    width = y_max - y_min

    mask = np.abs(y - y_center) > (width / 3.0)
    pos_cnt = np.sum(y > 0)
    if pos_cnt > 0.75 * len(y):
        mask_1 = mask & ((y - y_center) < 0).astype(bool)  # 靠近坐标轴的一侧
        mask_2 = mask & ((y - y_center) > 0).astype(bool)  # 远离坐标轴的一侧
        if abs(len(points[mask_1]) - len(points[mask_2])) > 200:
            return points[mask_1], 1, points[mask_2], 2
    else:
        mask_1 = mask & ((y - y_center) > 0).astype(bool)  # 靠近坐标轴的一侧
        mask_2 = mask & ((y - y_center) < 0).astype(bool)  # 远离坐标轴的一侧
    return points[mask_1], -1, points[mask_2], 1

def process_single_ship(points, nums=1, water_level=None):
    # ---------- taipu_main_side. x_center 判断 ----------
    x_min = points[:, 0].min()
    x_max = points[:, 0].max()
    x_center = (x_max + x_min) / 2
    if x_center < 50.0 or x_center > 135.0:
        return x_center, None, None, 0, 0, None, None

    # ---------- 2. 密度 ----------
    center_points = select_center(points)
    density = compute_average_density(center_points)
    if density == float('inf'):
        return x_center, density, None, 0, 0, None, None

    # ---------- 3. 粗糙度 ----------
    try:
        roughness = compute_roughness_entropy(center_points,
                                              radius=1.0 * (1 / density) ** (1/3))
    except:
        roughness = 1.0

    # ---------- 4. 旋转 ----------
    rotation_center_points, rotation_matrix, center = apply_rotation(
        points, center_points, density)

    # ---------- 5. y 方向分割 ----------
    n = max(1, int(len(rotation_center_points) / 400))

    y_1, direction_1, y_2, direction_2 = filter_by_y(rotation_center_points)

    # ---------- 6. 计算 freeboard_upper_bounds ----------
    freeboard_upper_bounds = []
    for y, direction in ((y_1, direction_1), (y_2, direction_2)):
        try:
            # 6.taipu_main_side 统计离群点过滤
            yoz_points = y.copy()
            yoz_points[:, 1] = 0
            nb_neighbors = 20 * density
            std_ratio = 0.15 * roughness
            mask = statistical_outlier(yoz_points,
                                       nb_neighbors=nb_neighbors,
                                       std_ratio=std_ratio,
                                       is_inliers=True)
            process_points = y[mask]

            # 6.2 分段
            divide_rotation_center_list = divide_points(process_points, axis=0, n=n)

            # 6.3 向量化提取最高点
            freeboard_upper_points = []
            for seg in divide_rotation_center_list:
                k = max(1, int(nums * roughness))
                y_coords = seg[:, 1] + 1000
                idx = np.argsort(direction * np.abs(y_coords))[:k]
                best = idx[np.argmax(seg[idx, 2])]
                freeboard_upper_points.append(seg[best])

            freeboard_upper_points = np.asarray(freeboard_upper_points)
            freeboard_upper_points = reverse_rotation(freeboard_upper_points,
                                                      rotation_matrix,
                                                      center)

            # 6.4 众数/中位数
            z_round = freeboard_upper_points[:, 2].round(2)
            freeboard_upper_bound = find_most_frequent_value(z_round, n=n)
            if freeboard_upper_bound is None:
                freeboard_upper_bound = np.median(z_round)
            freeboard_upper_bounds.append(freeboard_upper_bound)
        except:
            freeboard_upper_bounds.append(0)

    # ---------- 7. 下边界 + deck ----------
    if water_level is not None:
        freeboard_lower_bound = water_level
        deck = abs(freeboard_upper_bounds[0] - freeboard_lower_bound)
    else:
        freeboard_lower_bound = center_points[:, 2].min()
        deck = freeboard_upper_bounds[0] - freeboard_lower_bound

    return (x_center,
            density,
            roughness,
            freeboard_upper_bounds[0],
            freeboard_upper_bounds[1],
            freeboard_lower_bound,
            deck)


def remove_outliers_iqr(measurements, factor=1.5, variance_threshold=1e-3):
    """
    基于四分位数（IQR）方法去除测量值中的异常值。

    参数:
        measurements (numpy.ndarray): 测量值数组。
        factor (float): IQR的倍数，默认为1.5。

    返回:
        tuple:
            - filtered_measurements (numpy.ndarray): 去除异常值后的测量值数组。
            - outliers (numpy.ndarray): 被识别为异常值的数组。
    """
    # 确保 measurements 是 NumPy 数组
    measurements = np.array(measurements)
    # 将 None 值替换为 np.nan
    measurements = np.where(
        (measurements == None) | (measurements == 0),  # 用 | 替代 &
        np.nan,
        measurements
    )
    if np.var(measurements) < variance_threshold:
        # 过滤掉 NaN 值
        measurements = measurements[~np.isnan(measurements)]
        return measurements  # 方差太小，跳过 IQR 去噪

    # 将数组转换为数值类型
    try:
        measurements = pd.to_numeric(measurements, errors='coerce')
    except ValueError as e:
        raise ValueError("数据中包含无法转换为数值的值，请检查输入数据。") from e

    # 过滤掉 NaN 值
    measurements = measurements[~np.isnan(measurements)]

    return measurements

def split_filename_from_path(file_path):
    """
    从完整路径中提取文件名（不包括扩展名），并按 '_' 分隔为多个部分。
    参数: file_path (str): 完整的文件路径。
    返回: list: 文件名的各个部分。
    """
    # 提取文件名（包括扩展名）
    file_name_with_extension = os.path.basename(file_path)
    # 去掉扩展名
    file_name, _ = os.path.splitext(file_name_with_extension)
    # 使用 split() 方法按 '_' 分隔文件名
    parts = file_name.split('_')
    # 去除空字符串部分（如果存在）
    parts = [part for part in parts if part]

    return parts

def read_txt_dict(txt_data_path):
    # 获取所有txt文件
    file_names = glob.glob(f'{txt_data_path}/*.txt', recursive=True)
    for i, file_name in enumerate(file_names):
        print(f"[{i}]-{file_name}")
    file_idx = int(input("选择筛选的txt文件>>"))

    try:
        # 打开文件并逐行读取
        with open(file_names[file_idx], 'r', encoding='utf-8') as file:
            lines = file.readlines()  # 读取所有行

        # 用于存储解析后的字典
        txt_dicts = []

        # 遍历每一行，尝试解析为字典
        for line in lines:
            line = line.strip()  # 去掉首尾空格和换行符
            if not line:
                continue  # 跳过空行
            try:
                # 替换单引号为双引号（如果需要）
                line = line.replace("'", '"')
                # 使用 ast.literal_eval 安全地解析为字典
                txt_dict = ast.literal_eval(line)
                txt_dicts.append(txt_dict)  # 将解析后的字典添加到列表中
            except SyntaxError as e:
                print(f"解析错误：{e}，内容为：{line}")
            except Exception as e:
                print(f"其他错误：{e}，内容为：{line}")

    except FileNotFoundError:
        print(f"文件未找到，请检查路径：{file_names[file_idx]}")
    except Exception as e:
        print(f"发生错误：{e}")
    return txt_dicts

def find_lidar_interval_by_input(input_value, txt_dicts):
    # 遍历txt_dicts中的每一项
    for txt_item in txt_dicts:
        # 检查local_path是否包含输入值
        if input_value in txt_item.get("local_path", ""):
            # 返回lidar_interval对应的值
            return txt_item.get("lidar_interval")
    # 如果没有匹配到，返回None
    return None

def draw_steady_png(data, steady_value, channel_name, save_path):
    y = np.asarray(data, dtype=float)

    # 空数组 & 等值保护
    if y.size == 0:
        y = np.array([0.0])
    if np.all(y == y[0]):
        y = y + 1e-6 * np.random.randn(len(y))

    idx   = np.arange(len(y))
    total = len(y)

    # 稳态判定（最长区间）
    win   = max(5, min(30, len(y) // 4))
    roll  = pd.Series(y).rolling(win, min_periods=1).std()
    thresh = roll.quantile(0.25) if len(y) > win else np.inf
    steady_mask = roll < thresh

    # 提取所有区间，只保留最长一段
    starts, ends = [], []
    in_steady = False
    for i, val in enumerate(steady_mask):
        if val and not in_steady:
            starts.append(i)
            in_steady = True
        elif not val and in_steady:
            ends.append(i)
            in_steady = False
    if in_steady:
        ends.append(len(y))

    # 最长区间
    if starts:
        lengths = np.array(ends) - np.array(starts)
        max_idx = np.argmax(lengths)
        start, end = starts[max_idx], ends[max_idx]
        n_steady = end - start
    else:
        start, end, n_steady = None, None, 0

    # ±0.2 区间
    if steady_value is not None:
        mask_02 = (y >= steady_value - 0.2) & (y <= steady_value + 0.2)
        n_02    = mask_02.sum()
        closest_idx = np.argmin(np.abs(y - steady_value))
        closest_val = y[closest_idx]
    else:
        n_02 = 0
        closest_idx = None
        closest_val = None

    # 绘图
    fig, ax = plt.subplots(figsize=(10, 5))
    fig.subplots_adjust(left=0.06, right=0.75, top=0.9, bottom=0.15)

    ax.set_title(f'Extracting Steady-State Value from {channel_name}', fontsize=11, pad=8)
    ax.plot(idx, y, color='#0072bd', lw=1.2)
    ax.scatter(idx, y, color='#0072bd', s=12, zorder=3)

    # 仅画最长稳态条带
    if start is not None:
        ax.axvspan(start, end, color='#ff7f0e', alpha=0.25)

    if steady_value is not None:
        ax.axhspan(steady_value - 0.2, steady_value + 0.2,
                   color='yellow', alpha=0.25)
        ax.axhline(steady_value, color='red', ls='--', lw=1.5)
        ax.scatter(closest_idx, closest_val, color='darkgreen', s=50, zorder=5)

    # 竖排图例
    legend_texts = [
        'Measured data',
        f'Total Points : {total}',
        f'Steady Region : {n_steady}',
        f'Steady ±0.2 : {n_02}',
        f'Steady Value : {steady_value:.2f}' if steady_value is not None else 'Steady Value : N/A',
        f'Closest Point : {closest_val:.2f}' if closest_val is not None else 'Closest Point : N/A'
    ]
    ax.legend(legend_texts, loc='best', fontsize=8, frameon=False)

    y_min, y_max = y.min(), y.max()
    if y_min == y_max:
        y_min, y_max = y_min - 1, y_max + 1
    ax.set_xlim(0, max(1, len(y)))
    ax.set_ylim(y_min - 0.05 * (y_max - y_min), y_max + 0.05 * (y_max - y_min))

    ax.set_xlabel('Index', fontsize=9)
    ax.set_ylabel('Value', fontsize=9)
    ax.grid(True, which='both', ls='--', lw=0.5, alpha=0.3)
    if os.path.exists(save_path):
        try:
            os.remove(save_path)
        except PermissionError:
            print(f"文件被占用：{save_path}")

    plt.savefig(save_path, dpi=200, bbox_inches='tight')
    plt.close(fig)

if __name__ == '__main__':
    o3d.utility.random.seed(42)
    directory_path    = r"D:\li3D-ML\data\WuHanQingShanDaQiao\data_tilt"
    output_excel_path = './table/output_double2.xlsx'
    png_dir_path = './table'


    # 断点续跑（仅按 Excel 最后一行）
    if os.path.exists(output_excel_path):
        last = pd.read_excel(output_excel_path).iloc[-1]
        last_dir, last_subdir = last['Directory'], last['Subdirectory']
    else:
        last_dir = last_subdir = None

    dirs = [d for d in os.listdir(directory_path)
            if os.path.isdir(os.path.join(directory_path, d))]
    start_index = dirs.index(last_dir) + 1 if last_dir else 0

    for dir in dirs[start_index:]:
        detail_dir = os.path.join(directory_path, dir, "detail")
        if not os.path.isdir(detail_dir):
            continue
        sub_dirs = [s for s in os.listdir(detail_dir)
                    if os.path.isdir(os.path.join(detail_dir, s))]

        start_sub = sub_dirs.index(last_subdir) + 1 if dir == last_dir and last_subdir else 0

        for sub_dir in sub_dirs[start_sub:1]:
            # 跳过直到 (last_dir, last_subdir) 之后
            if last_dir is not None and last_subdir is not None:
                if dir < last_dir or (dir == last_dir and sub_dir <= last_subdir):
                    continue

            current_path = os.path.join(detail_dir, sub_dir)
            pcd_files = glob.glob(os.path.join(current_path, '*.pcd'))

            frame_freeboard_upper_bound_1 = []
            frame_freeboard_upper_bound_2 = []
            loop = tqdm(pcd_files, desc=f"{dir}/{sub_dir}")
            for pcd_file in loop:
                l = split_filename_from_path(pcd_file)
                pcd = o3d.io.read_point_cloud(pcd_file, format="pcd")
                (x_center, density, roughness,
                 freeboard_upper_bound_1,
                 freeboard_upper_bound_2,
                 _, _) = process_single_ship(
                    np.asarray(pcd.points)[:int(l[-1])])
                loop.set_description(f'{dir}')
                loop.set_postfix(x_center=x_center, density=density, roughness=roughness)
                if freeboard_upper_bound_1 is not None:
                    frame_freeboard_upper_bound_1.append(freeboard_upper_bound_1)
                if freeboard_upper_bound_2 is not None:
                    frame_freeboard_upper_bound_2.append(freeboard_upper_bound_2)

                # ---------- 计算高斯峰值 ----------
            if not frame_freeboard_upper_bound_1:
                steady_state_freeboard_upper_bound_1 = None
            else:
                steady_state_freeboard_upper_bound_1 = gaussian_approximation(
                    remove_outliers_iqr(frame_freeboard_upper_bound_1))
            if not frame_freeboard_upper_bound_2:
                steady_state_freeboard_upper_bound_2 = None
            else:
                steady_state_freeboard_upper_bound_2 = gaussian_approximation(
                    remove_outliers_iqr(frame_freeboard_upper_bound_2))

                # 3. 生成图片并计算相对路径
                png_sub_dir = os.path.join(png_dir_path, '.steady_double', dir)
                os.makedirs(png_sub_dir, exist_ok=True)

                png_path_1 = os.path.join(png_sub_dir , 'steady_1.png')
                png_path_2 = os.path.join(png_sub_dir , 'steady_2.png')

                draw_steady_png(frame_freeboard_upper_bound_1,
                                steady_state_freeboard_upper_bound_1,
                                dir,
                                png_path_1)

                draw_steady_png(frame_freeboard_upper_bound_2,
                                steady_state_freeboard_upper_bound_2,
                                dir,
                                png_path_2)

                # 4. 计算相对于 Excel 文件的相对路径
                rel_1 = os.path.relpath(png_path_1, start=png_dir_path)
                rel_2 = os.path.relpath(png_path_2, start=png_dir_path)

                # 5. 组装 DataFrame 并写入 Excel
                data = [{
                    'Directory': dir,
                    'Subdirectory': 'sub_dir_placeholder',
                    'Steady State Freeboard Upper Bound taipu_main_side': steady_state_freeboard_upper_bound_1,
                    'Steady State Freeboard Upper Bound 2': steady_state_freeboard_upper_bound_2,
                    'Steady State Plot taipu_main_side': f'=HYPERLINK("{rel_1}", "Plot_1")',
                    'Steady State Plot 2': f'=HYPERLINK("{rel_2}", "Plot_2")'
                }]
            df = pd.DataFrame(data)

            if os.path.exists(output_excel_path):
                wb = load_workbook(output_excel_path)
                ws = wb.active
                next_row = ws.max_row + 1
                for col, val in enumerate(df.iloc[0], 1):
                    ws.cell(row=next_row, column=col, value=val)
                wb.save(output_excel_path)
            else:
                df.to_excel(output_excel_path, index=False)

            print(f"mask_1峰值: {steady_state_freeboard_upper_bound_1}")
            print(f"mask_2峰值: {steady_state_freeboard_upper_bound_2}")

    print(f"全部完成，结果见 {output_excel_path}")